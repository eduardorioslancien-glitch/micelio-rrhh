# -*- coding: utf-8 -*-
"""Exporta TODOS los trabajadores/postulantes registrados en el portal a un
único Excel multi-hoja (punto 6 del pedido del usuario), con el mismo
lenguaje visual de la Base de Datos Maestra del Kit RR.HH. DIGETEL GROUP:
  - "Datos Generales": una fila por persona con los campos planos de la ficha.
  - "Familia", "Educación", "Experiencia", "Capacitaciones": una fila por
    cada elemento de esas tablas relacionadas (varias filas por persona).
  - "Documentos Adjuntos": una fila por archivo subido (CV, CUL, etc.).
  - "Estado de Firmas": estado de cada documento del legajo por persona.
"""
import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import Employee, DOC_TYPES, ATTACHMENT_TYPES

NAVY = "1246AB"
LIGHT2 = "EEF2FA"
GRAY = "595959"
WHITE = "FFFFFF"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXPORT_PATH = os.path.join(BASE_DIR, "generated", "export_portal_rrhh.xlsx")

ATTACHMENT_LABELS = dict(ATTACHMENT_TYPES)

# Columnas de la hoja "Datos Generales": (encabezado, función sobre Employee)
GENERAL_COLUMNS = [
    ("Código", lambda e: f"E-{e.id:04d}"),
    ("Nombre Completo", lambda e: e.nombre_completo),
    ("Empresa", lambda e: e.empresa),
    ("Correo Personal", lambda e: (e.ficha_data or {}).get("correo_personal", "") or e.email or ""),
    ("Correo Corporativo", lambda e: (e.ficha_data or {}).get("correo_corporativo", "")),
    ("Tipo Doc.", lambda e: (e.ficha_data or {}).get("tipo_documento", "")),
    ("N° Documento", lambda e: (e.ficha_data or {}).get("numero_documento", "")),
    ("Sexo", lambda e: (e.ficha_data or {}).get("sexo", "")),
    ("Estado Civil", lambda e: (e.ficha_data or {}).get("estado_civil", "")),
    ("Fecha Nacimiento", lambda e: (e.ficha_data or {}).get("fecha_nacimiento", "")),
    ("Edad", lambda e: (e.ficha_data or {}).get("edad", "")),
    ("Dirección", lambda e: (e.ficha_data or {}).get("direccion", "")),
    ("Distrito", lambda e: (e.ficha_data or {}).get("distrito", "")),
    ("Provincia", lambda e: (e.ficha_data or {}).get("provincia", "")),
    ("Departamento", lambda e: (e.ficha_data or {}).get("departamento", "")),
    ("Celular", lambda e: (e.ficha_data or {}).get("celular", "")),
    ("Área", lambda e: (e.ficha_data or {}).get("area", "")),
    ("Gerencia", lambda e: (e.ficha_data or {}).get("gerencia", "")),
    ("Cargo", lambda e: (e.ficha_data or {}).get("cargo", "")),
    ("Sede", lambda e: (e.ficha_data or {}).get("sede", "")),
    ("Centro de Costos", lambda e: (e.ficha_data or {}).get("centro_costos", "")),
    ("Jefe Inmediato", lambda e: (e.ficha_data or {}).get("jefe_inmediato", "")),
    ("Tipo de Contrato", lambda e: (e.ficha_data or {}).get("tipo_contrato", "")),
    ("Fecha de Ingreso", lambda e: (e.ficha_data or {}).get("fecha_ingreso", "")),
    ("Modalidad", lambda e: (e.ficha_data or {}).get("modalidad", "")),
    ("Turno", lambda e: (e.ficha_data or {}).get("turno", "")),
    ("Remuneración", lambda e: (e.ficha_data or {}).get("remuneracion", "")),
    ("Banco (Haberes)", lambda e: (e.ficha_data or {}).get("banco_haberes", "")),
    ("Cuenta (Haberes)", lambda e: (e.ficha_data or {}).get("cuenta_haberes", "")),
    ("CCI (Haberes)", lambda e: (e.ficha_data or {}).get("cci_haberes", "")),
    ("Banco CTS", lambda e: (e.ficha_data or {}).get("banco_cts", "")),
    ("Cuenta CTS", lambda e: (e.ficha_data or {}).get("cuenta_cts", "")),
    ("Sistema Pensión", lambda e: (e.ficha_data or {}).get("sistema_pension", "")),
    ("AFP", lambda e: (e.ficha_data or {}).get("afp", "")),
    ("CUSPP", lambda e: (e.ficha_data or {}).get("cuspp", "")),
    ("Grupo Sanguíneo", lambda e: (e.ficha_data or {}).get("grupo_sanguineo", "")),
    ("EPS", lambda e: (e.ficha_data or {}).get("eps", "")),
    ("EsSalud", lambda e: (e.ficha_data or {}).get("essalud", "")),
    ("Fecha de Creación del Enlace", lambda e: e.created_at.strftime("%d/%m/%Y %H:%M") if e.created_at else ""),
    ("Fecha de Apertura", lambda e: e.link_opened_at.strftime("%d/%m/%Y %H:%M") if e.link_opened_at else "Sin abrir"),
    ("Fecha de Legajo Completo", lambda e: e.completed_at.strftime("%d/%m/%Y %H:%M") if e.completed_at else "Pendiente"),
]


def _style_header(ws, row, ncols, title=None):
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws["A1"].font = Font(bold=True, size=13, color=NAVY)
        ws["A1"] = title
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = Font(bold=True, color=WHITE, size=9)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate


def build_export(db, employees=None):
    if employees is None:
        employees = db.query(Employee).order_by(Employee.created_at.desc()).all()
    wb = Workbook()
    today = datetime.date.today().strftime("%d/%m/%Y")

    # ---- Hoja 1: Datos Generales ----
    ws = wb.active
    ws.title = "Datos Generales"
    headers = [c[0] for c in GENERAL_COLUMNS]
    ws.append([f"BASE DE DATOS — PORTAL RR.HH. DIGETEL GROUP (exportado {today})"])
    ws.append(headers)
    _style_header(ws, 2, len(headers), title=ws["A1"].value)
    for e in employees:
        ws.append([fn(e) for _, fn in GENERAL_COLUMNS])
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"

    # ---- Hoja 2: Estado de Firmas ----
    ws2 = wb.create_sheet("Estado de Firmas")
    headers2 = ["Código", "Nombre Completo", "Empresa"] + [lbl for _, lbl in DOC_TYPES]
    ws2.append(headers2)
    _style_header(ws2, 1, len(headers2))
    status_lbl = {"pendiente": "Pendiente", "abierto": "En proceso", "firmado": "Firmado"}
    for e in employees:
        docs = {d.doc_type: d.status for d in e.documents}
        ws2.append([f"E-{e.id:04d}", e.nombre_completo, e.empresa] +
                   [status_lbl.get(docs.get(k, "pendiente"), "Pendiente") for k, _ in DOC_TYPES])
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}1"

    # ---- Hoja 3: Familia ----
    ws3 = wb.create_sheet("Familia")
    headers3 = ["Código", "Nombre Trabajador", "Parentesco", "Nombre Familiar", "DNI",
                "Fecha Nacimiento", "Depende Económicamente", "Derechohabiente EsSalud"]
    ws3.append(headers3)
    _style_header(ws3, 1, len(headers3))
    for e in employees:
        for f in (e.familia_data or []):
            ws3.append([f"E-{e.id:04d}", e.nombre_completo, f.get("parentesco", ""), f.get("nombre", ""),
                        f.get("dni", ""), f.get("fecha_nacimiento", ""),
                        "Sí" if f.get("depende_economicamente") else "No",
                        "Sí" if f.get("derechohabiente_essalud") else "No"])

    # ---- Hoja 4: Educación ----
    ws4 = wb.create_sheet("Educación")
    headers4 = ["Código", "Nombre Trabajador", "Institución", "Carrera", "Nivel", "Grado", "Año", "Estado"]
    ws4.append(headers4)
    _style_header(ws4, 1, len(headers4))
    for e in employees:
        for ed in (e.educacion_data or []):
            ws4.append([f"E-{e.id:04d}", e.nombre_completo, ed.get("institucion", ""), ed.get("carrera", ""),
                        ed.get("nivel", ""), ed.get("grado", ""), ed.get("anio", ""), ed.get("estado", "")])

    # ---- Hoja 5: Experiencia ----
    ws5 = wb.create_sheet("Experiencia")
    headers5 = ["Código", "Nombre Trabajador", "Empresa", "Cargo", "Periodo", "Funciones"]
    ws5.append(headers5)
    _style_header(ws5, 1, len(headers5))
    for e in employees:
        for ex in (e.experiencia_data or []):
            ws5.append([f"E-{e.id:04d}", e.nombre_completo, ex.get("empresa", ""), ex.get("cargo", ""),
                        ex.get("periodo", ""), ex.get("funciones", "")])

    # ---- Hoja 6: Capacitaciones ----
    ws6 = wb.create_sheet("Capacitaciones")
    headers6 = ["Código", "Nombre Trabajador", "Curso", "Institución", "Horas", "Año"]
    ws6.append(headers6)
    _style_header(ws6, 1, len(headers6))
    for e in employees:
        for cap in (e.capacitaciones_data or []):
            ws6.append([f"E-{e.id:04d}", e.nombre_completo, cap.get("curso", ""), cap.get("institucion", ""),
                        cap.get("horas", ""), cap.get("anio", "")])

    # ---- Hoja 7: Documentos Adjuntos ----
    ws7 = wb.create_sheet("Documentos Adjuntos")
    headers7 = ["Código", "Nombre Trabajador", "Tipo de Documento", "Archivo", "Fecha de Carga"]
    ws7.append(headers7)
    _style_header(ws7, 1, len(headers7))
    for e in employees:
        for att in (e.attachments or []):
            ws7.append([f"E-{e.id:04d}", e.nombre_completo, ATTACHMENT_LABELS.get(att.tipo, att.tipo),
                        att.filename, att.uploaded_at.strftime("%d/%m/%Y %H:%M") if att.uploaded_at else ""])

    os.makedirs(os.path.dirname(EXPORT_PATH), exist_ok=True)
    wb.save(EXPORT_PATH)
    return EXPORT_PATH
